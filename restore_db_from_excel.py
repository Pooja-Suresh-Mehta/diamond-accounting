"""
Restore diamond_accounting.db from a backup Excel file.

Usage:
  source backend/venv/bin/activate
  python3 restore_db_from_excel.py

The script will:
  - DELETE all rows from every table (keeps schema intact)
  - Re-import every sheet from the backup Excel
"""

import sqlite3
import openpyxl
from datetime import datetime

import sys

DB_PATH   = "backend/diamond_accounting.db"
BACKUP_XL = sys.argv[1] if len(sys.argv) > 1 else "backend/poojan_gems_backup_20260428_115346.xlsx"


def sheet_to_rows(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return headers, rows


def fmt_val(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return v


def insert_rows(conn, table, headers, rows):
    if not rows:
        return 0
    c = conn.cursor()
    c.execute(f"PRAGMA table_info({table})")
    db_cols = {r[1] for r in c.fetchall()}
    cols = [h for h in headers if h in db_cols]
    placeholders = ", ".join("?" * len(cols))
    sql = f"INSERT OR REPLACE INTO {table} ({', '.join(cols)}) VALUES ({placeholders})"
    data = [[fmt_val(row.get(col)) for col in cols] for row in rows]
    c.executemany(sql, data)
    return len(data)


conn = sqlite3.connect(DB_PATH)
conn.execute("PRAGMA foreign_keys = OFF")

print("Loading backup Excel…")
wb = openpyxl.load_workbook(BACKUP_XL)

c = conn.cursor()
c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
db_tables = [r[0] for r in c.fetchall()]

print(f"Clearing {len(db_tables)} tables…")
for t in db_tables:
    conn.execute(f"DELETE FROM {t}")
conn.commit()

print("\nRestoring:")
for sheet_name in wb.sheetnames:
    if sheet_name not in db_tables:
        print(f"  SKIP  {sheet_name} (not in DB schema)")
        continue
    ws = wb[sheet_name]
    headers, rows = sheet_to_rows(ws)
    n = insert_rows(conn, sheet_name, headers, rows)
    print(f"  {sheet_name:40s}  {n} rows")

conn.commit()

print("\nVerification:")
for t in db_tables:
    c.execute(f"SELECT COUNT(*) FROM {t}")
    cnt = c.fetchone()[0]
    if cnt:
        print(f"  {t:40s}  {cnt}")

conn.execute("PRAGMA foreign_keys = ON")

# Sanity check: last 3 rows of parcel_masters in Excel vs DB (matched by id)
print("\nSanity check — last 3 parcel_masters rows:")
ws_pm = wb["parcel_masters"]
xl_headers, xl_rows = sheet_to_rows(ws_pm)
ok = True
for row in xl_rows[-3:]:
    row_id = row.get("id")
    lot = row.get("lot_no")
    item = row.get("item_name")
    ow = row.get("opening_weight_carats")
    c.execute("SELECT lot_no, item_name, opening_weight_carats FROM parcel_masters WHERE id=?", (row_id,))
    db_row = c.fetchone()
    if db_row and db_row[0] == str(lot) and db_row[1] == item:
        print(f"  OK  lot={lot}  item={item}  opening_wt={ow}")
    else:
        ok = False
        print(f"  MISMATCH  Excel: lot={lot} item={item}  |  DB: {db_row}")
if ok:
    print("  All rows match.")

conn.close()
print("\nDone. DB restored from", BACKUP_XL)
