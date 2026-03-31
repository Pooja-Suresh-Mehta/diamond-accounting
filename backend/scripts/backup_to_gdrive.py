"""
Backup all SQLite tables to a single Excel file, then upload to Google Drive.

Requirements (add to requirements.txt or pip install separately):
    google-api-python-client
    google-auth-httplib2
    google-auth-oauthlib
    openpyxl  (already in requirements.txt)

Setup:
    1. Go to https://console.cloud.google.com → New Project
    2. Enable "Google Drive API"
    3. Create credentials → Service Account → download JSON key
    4. Share the target Drive folder with the service account email
    5. Set env vars:
         GDRIVE_SERVICE_ACCOUNT_JSON=/path/to/service_account.json
         GDRIVE_FOLDER_ID=<folder_id_from_drive_url>

Usage:
    cd backend
    GDRIVE_SERVICE_ACCOUNT_JSON=./service_account.json \
    GDRIVE_FOLDER_ID=1AbCdEfGhI... \
    python scripts/backup_to_gdrive.py
"""
import os
import io
import sqlite3
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH = Path(__file__).parent.parent / "poojan_gems.db"
SERVICE_ACCOUNT_JSON = os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON")
GDRIVE_FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID")


def export_all_tables_to_excel() -> bytes:
    """Export every table in the SQLite DB to a separate Excel sheet."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()

    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [row[0] for row in cur.fetchall() if not row[0].startswith("sqlite_")]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")

    for table in tables:
        cur.execute(f"PRAGMA table_info({table})")
        columns = [col[1] for col in cur.fetchall()]

        cur.execute(f"SELECT * FROM {table}")
        rows = cur.fetchall()

        # Sheet names max 31 chars
        ws = wb.create_sheet(title=table[:31])

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_to_gdrive(file_bytes: bytes, filename: str) -> str:
    """Upload bytes to Google Drive and return the file URL."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    if not SERVICE_ACCOUNT_JSON:
        raise EnvironmentError("GDRIVE_SERVICE_ACCOUNT_JSON not set")
    if not GDRIVE_FOLDER_ID:
        raise EnvironmentError("GDRIVE_FOLDER_ID not set")

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON,
        scopes=["https://www.googleapis.com/auth/drive.file"],
    )
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    file_metadata = {
        "name": filename,
        "parents": [GDRIVE_FOLDER_ID],
    }
    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    result = service.files().create(
        body=file_metadata, media_body=media, fields="id,webViewLink"
    ).execute()

    return result.get("webViewLink", result["id"])


def main():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"poojan_gems_backup_{timestamp}.xlsx"

    print(f"Exporting all tables to {filename}...")
    file_bytes = export_all_tables_to_excel()
    print(f"  Excel size: {len(file_bytes):,} bytes")

    if SERVICE_ACCOUNT_JSON and GDRIVE_FOLDER_ID:
        print("Uploading to Google Drive...")
        url = upload_to_gdrive(file_bytes, filename)
        print(f"  Uploaded: {url}")
    else:
        # Save locally if Drive not configured
        out_path = Path(__file__).parent.parent / filename
        out_path.write_bytes(file_bytes)
        print(f"  Saved locally (Drive not configured): {out_path}")
        print("  Set GDRIVE_SERVICE_ACCOUNT_JSON and GDRIVE_FOLDER_ID to upload to Drive.")


if __name__ == "__main__":
    main()
