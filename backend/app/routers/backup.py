"""Backup & Restore: export all tables to Excel, import from Excel."""
import io
import os
import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import text, inspect as sa_inspect
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db, engine
from app.auth import get_current_user

try:
    import openpyxl
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/api/backup", tags=["backup"])

# Tables to skip (internal / alembic)
_SKIP_TABLES = {"alembic_version"}


async def _get_table_names():
    """Return all user table names in the DB."""
    async with engine.connect() as conn:
        def _inspect(sync_conn):
            insp = sa_inspect(sync_conn)
            return insp.get_table_names()
        return await conn.run_sync(_inspect)


@router.get("/export")
async def export_backup(user=Depends(get_current_user)):
    """Export entire database as a single .xlsx file (one sheet per table)."""
    if not openpyxl:
        raise HTTPException(500, "openpyxl not installed on server")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    table_names = await _get_table_names()

    async with engine.connect() as conn:
        for tbl in sorted(table_names):
            if tbl in _SKIP_TABLES:
                continue
            rows = (await conn.execute(text(f'SELECT * FROM "{tbl}"'))).fetchall()
            keys_result = await conn.execute(text(f'SELECT * FROM "{tbl}" LIMIT 0'))
            col_names = list(keys_result.keys())

            ws = wb.create_sheet(title=tbl[:31])  # Excel sheet name max 31 chars
            ws.append(col_names)
            for row in rows:
                ws.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"poojan_gems_backup_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_backup(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Import an Excel backup: replaces ALL data in every sheet's matching table."""
    if not openpyxl:
        raise HTTPException(500, "openpyxl not installed on server")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx files are accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))

    existing_tables = set(await _get_table_names())
    imported = []
    skipped = []

    async with engine.begin() as conn:
        for sheet_name in wb.sheetnames:
            tbl = sheet_name.strip()
            if tbl not in existing_tables:
                skipped.append(tbl)
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 1:
                skipped.append(tbl)
                continue

            headers = [str(h).strip() for h in rows[0]]
            data_rows = rows[1:]

            # Clear existing data
            await conn.execute(text(f'DELETE FROM "{tbl}"'))

            if data_rows:
                placeholders = ", ".join([f":{h}" for h in headers])
                insert_sql = f'INSERT INTO "{tbl}" ({", ".join(headers)}) VALUES ({placeholders})'
                for row in data_rows:
                    params = {}
                    for i, h in enumerate(headers):
                        val = row[i] if i < len(row) else None
                        params[h] = val
                    await conn.execute(text(insert_sql), params)

            imported.append(tbl)

    return {
        "status": "ok",
        "imported_tables": imported,
        "skipped_sheets": skipped,
        "message": f"Restored {len(imported)} tables from backup.",
    }


@router.get("/db-copy")
async def download_db_copy(user=Depends(get_current_user)):
    """Download a raw copy of the SQLite database file."""
    from app.config import get_settings
    settings = get_settings()
    db_url = settings.DATABASE_URL

    if "sqlite" not in db_url:
        raise HTTPException(400, "Raw DB download only available for SQLite")

    # Extract path from sqlite URL
    db_path = db_url.split("///")[-1]
    if not os.path.exists(db_path):
        raise HTTPException(404, "Database file not found")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def iter_file():
        with open(db_path, "rb") as f:
            while chunk := f.read(8192):
                yield chunk

    return StreamingResponse(
        iter_file(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="poojan_gems_db_{timestamp}.sqlite"'},
    )


@router.get("/info")
async def backup_info(user=Depends(get_current_user)):
    """Return table names and row counts for backup preview."""
    table_names = await _get_table_names()
    info = []
    async with engine.connect() as conn:
        for tbl in sorted(table_names):
            if tbl in _SKIP_TABLES:
                continue
            result = await conn.execute(text(f'SELECT COUNT(*) FROM "{tbl}"'))
            count = result.scalar()
            info.append({"table": tbl, "rows": count})
    return {"tables": info}
