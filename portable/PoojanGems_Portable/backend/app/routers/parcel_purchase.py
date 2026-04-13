from datetime import date, datetime
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth import get_current_user
from app.database import get_db
from app.models.models import AccountMaster, ParcelMaster, ParcelPurchase, ParcelPurchaseItem, User
from app.schemas import ParcelPurchaseCreate, ParcelPurchaseOut, ParcelPurchaseUpdate
from app.utils import (
    CATEGORIES, CURRENCIES, CURRENCY_RATES, PAYMENT_STATUSES, PURCHASE_TYPES, SUB_TYPES,
    adjust_parcel_stock, ensure_unique, get_actor_name, next_number,
    parse_date_value, parse_float_value, post_ledger_entries, reverse_ledger_entries,
)

router = APIRouter(prefix="/api/parcel/purchase", tags=["parcel-purchase"])


def _calc_totals(p: ParcelPurchase):
    total_carats = sum((i.selected_carat or i.issue_carats or 0) for i in p.items)
    total_amount = sum((i.amount or 0) for i in p.items)
    p.total_carats = float(total_carats or 0)
    p.total_amount = float(total_amount or 0)
    p.inr_amt = float(p.inr_final_amount or p.transaction_final_amount or total_amount or 0)
    p.usd_amt = float(p.usd_final_amount or total_amount or 0)


def _norm_header(v: str) -> str:
    return "".join(ch for ch in (v or "").strip().lower() if ch.isalnum())


async def _upsert_parcel_items(
    db: AsyncSession,
    *,
    company_id: str,
    created_by_name: str,
    items: list,
    inr_rate: float = 85,
):
    def _apply_first_purchase_costs(row: ParcelMaster, item):
        if float(row.purchase_cost_usd_carat or 0) > 0:
            return
        issue_carats = float(item.issue_carats or 0)
        rate_usd_per_carat = float(item.rate or 0)
        if issue_carats <= 0 or rate_usd_per_carat <= 0:
            return
        selected_inr_rate = float(inr_rate or 0)
        purchase_cost_usd_amount = rate_usd_per_carat * issue_carats
        purchase_cost_inr_amount = purchase_cost_usd_amount * selected_inr_rate
        asking_price_usd_carats = selected_inr_rate * issue_carats
        asking_usd_amount = asking_price_usd_carats * issue_carats
        asking_price_inr_carats = asking_price_usd_carats * selected_inr_rate

        row.usd_to_inr_rate = selected_inr_rate
        row.purchase_cost_usd_carat = rate_usd_per_carat
        row.purchase_cost_usd_amount = purchase_cost_usd_amount
        row.purchase_cost_inr_carat = selected_inr_rate
        row.purchase_cost_inr_amount = purchase_cost_inr_amount
        row.asking_price_usd_carats = asking_price_usd_carats
        row.asking_usd_amount = asking_usd_amount
        row.asking_price_inr_carats = asking_price_inr_carats
        row.asking_inr_amount = asking_usd_amount * selected_inr_rate

    lot_numbers = [str(i.lot_number).strip() for i in items if getattr(i, "lot_number", None) and str(i.lot_number).strip()]
    if not lot_numbers:
        return
    existing_rows = (await db.execute(
        select(ParcelMaster).where(
            ParcelMaster.company_id == company_id,
            func.lower(ParcelMaster.lot_no).in_([ln.lower() for ln in lot_numbers]),
        )
    )).scalars().all()
    existing = {r.lot_no.strip().lower(): r for r in existing_rows}

    for item in items:
        lot_no = (item.lot_number or "").strip()
        if not lot_no:
            continue
        row = existing.get(lot_no.lower())
        if row:
            row.item_name = item.item_name or row.item_name
            row.shape = item.shape or row.shape
            row.color = item.color or row.color
            row.clarity = item.clarity or row.clarity
            row.size = item.size or row.size
            row.sieve_mm = item.sieve or row.sieve_mm
            row.opening_weight_carats = float(item.selected_carat or item.issue_carats or row.opening_weight_carats or 0)
            row.asking_inr_amount = float(item.amount or row.asking_inr_amount or 0)
            _apply_first_purchase_costs(row, item)
            continue

        if not (item.shape and item.size):
            continue
        new_row = ParcelMaster(
            company_id=company_id,
            lot_no=lot_no,
            item_name=item.item_name or lot_no,
            shape=item.shape,
            color=item.color,
            clarity=item.clarity,
            size=item.size,
            sieve_mm=item.sieve,
            opening_weight_carats=float(item.selected_carat or item.issue_carats or 0),
            asking_inr_amount=float(item.amount or 0),
            created_by_name=created_by_name,
        )
        _apply_first_purchase_costs(new_row, item)
        db.add(new_row)


# ── Fixed route ordering: static paths BEFORE /{id} ─────

@router.get("/options")
async def get_purchase_options(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    CUSTOMER_TYPES = ["customer", "overseas customer", "individual", "supplier", "overseas supplier"]
    parties = (await db.execute(
        select(AccountMaster.account_group_name)
        .where(
            AccountMaster.company_id == current_user.company_id,
            func.lower(AccountMaster.account_type).in_(CUSTOMER_TYPES),
        )
        .order_by(AccountMaster.account_group_name)
    )).scalars().all()
    brokers = (await db.execute(
        select(AccountMaster.account_group_name)
        .where(AccountMaster.company_id == current_user.company_id, func.lower(AccountMaster.account_type) == "broker")
        .order_by(AccountMaster.account_group_name)
    )).scalars().all()
    parcel_rows = (await db.execute(
        select(ParcelMaster).where(ParcelMaster.company_id == current_user.company_id).order_by(ParcelMaster.lot_no)
    )).scalars().all()
    next_inv = await next_number(db, ParcelPurchase, ParcelPurchase.invoice_number, current_user.company_id)
    return {
        "types": PURCHASE_TYPES,
        "sub_types": SUB_TYPES,
        "categories": CATEGORIES,
        "currencies": CURRENCIES,
        "currency_rates": CURRENCY_RATES,
        "parties": parties,
        "brokers": brokers,
        "lot_numbers": [r.lot_no for r in parcel_rows if r.lot_no],
        "lot_items": [{
            "lot_no": r.lot_no, "item_name": r.item_name, "shape": r.shape,
            "color": r.color, "clarity": r.clarity, "size": r.size, "sieve": r.sieve_mm,
            "opening_weight_carats": r.opening_weight_carats,
            "amount": r.asking_inr_amount,
        } for r in parcel_rows if r.lot_no],
        "payment_statuses": PAYMENT_STATUSES,
        "next_invoice_number": next_inv,
    }


@router.get("/export")
async def export_purchases(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(ParcelPurchase).where(ParcelPurchase.company_id == current_user.company_id).order_by(ParcelPurchase.created_at.desc())
    )).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase"
    ws.append(["Invoice", "Date", "Type", "Sub Type", "Category", "Party", "Carats", "Amount", "Currency", "INR Amt", "USD Amt", "DueDate", "Payment Status"])
    for r in rows:
        ws.append([
            r.invoice_number, r.date.isoformat() if r.date else "", r.purchase_type,
            r.sub_type or "", r.category or "", r.party or "", float(r.total_carats or 0),
            float(r.total_amount or 0), r.currency or "", float(r.inr_amt or 0),
            float(r.usd_amt or 0), r.due_date.isoformat() if r.due_date else "", r.payment_status or "",
        ])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase.xlsx"},
    )


@router.get("/template")
async def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "PurchaseTemplate"
    ws.append(["Invoice", "Date(YYYY-MM-DD)", "Type", "Sub Type", "Category", "Party", "Carats", "Amount", "Currency", "INR Amt", "USD Amt", "DueDate(YYYY-MM-DD)", "Payment Status"])
    ws.append(["INV-001", "2026-03-12", "LOCAL", "DIRECT", "Natural Diamond", "Party A", 1.5, 1200, "USD", 102000, 1200, "2026-03-20", "Pending"])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_template.xlsx"},
    )


@router.get("", response_model=list[ParcelPurchaseOut])
async def list_purchases(
    search: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(ParcelPurchase).options(selectinload(ParcelPurchase.items)).where(ParcelPurchase.company_id == current_user.company_id)
    if search:
        like = f"%{search.strip()}%"
        q = q.where(
            ParcelPurchase.invoice_number.ilike(like) |
            ParcelPurchase.bill_no.ilike(like) |
            ParcelPurchase.party.ilike(like)
        )
    q = q.order_by(ParcelPurchase.created_at.desc())
    q = q.offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(q)).scalars().all()
    return [ParcelPurchaseOut.model_validate(r) for r in rows]


@router.get("/{purchase_id}", response_model=ParcelPurchaseOut)
async def get_purchase(
    purchase_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(
        select(ParcelPurchase).options(selectinload(ParcelPurchase.items))
        .where(ParcelPurchase.id == str(purchase_id), ParcelPurchase.company_id == current_user.company_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase not found")
    return ParcelPurchaseOut.model_validate(row)


@router.post("", response_model=ParcelPurchaseOut, status_code=status.HTTP_201_CREATED)
async def create_purchase(
    payload: ParcelPurchaseCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await ensure_unique(db, ParcelPurchase, ParcelPurchase.invoice_number, current_user.company_id, payload.invoice_number, label="Invoice Number")
    data = payload.model_dump(exclude={"items"})
    row = ParcelPurchase(company_id=current_user.company_id, **data)
    row.created_by_name = get_actor_name(current_user)
    row.items = [ParcelPurchaseItem(**item.model_dump()) for item in payload.items]
    _calc_totals(row)
    db.add(row)

    await _upsert_parcel_items(db, company_id=current_user.company_id, created_by_name=get_actor_name(current_user), items=payload.items, inr_rate=payload.inr_rate)
    await adjust_parcel_stock(db, company_id=current_user.company_id, items=payload.items, operation="purchase")

    await db.flush()
    # Ledger: Debit Purchase, Credit Supplier
    amount = float(row.transaction_final_amount or row.total_amount or 0)
    party = row.party or "Unknown Supplier"
    if amount > 0:
        await post_ledger_entries(db, company_id=current_user.company_id, transaction_type="purchase", transaction_id=row.id, transaction_date=row.date, created_by=get_actor_name(current_user), entries=[
            {"account_name": "Purchase", "debit": amount, "credit": 0, "narration": f"Purchase {row.invoice_number} from {party}"},
            {"account_name": party, "debit": 0, "credit": amount, "narration": f"Purchase {row.invoice_number}"},
        ])

    await db.commit()
    await db.refresh(row)
    row = (await db.execute(select(ParcelPurchase).options(selectinload(ParcelPurchase.items)).where(ParcelPurchase.id == row.id))).scalar_one()
    return ParcelPurchaseOut.model_validate(row)


@router.put("/{purchase_id}", response_model=ParcelPurchaseOut)
async def update_purchase(
    purchase_id: UUID,
    payload: ParcelPurchaseUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(
        select(ParcelPurchase).options(selectinload(ParcelPurchase.items))
        .where(ParcelPurchase.id == str(purchase_id), ParcelPurchase.company_id == current_user.company_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase not found")

    await ensure_unique(db, ParcelPurchase, ParcelPurchase.invoice_number, current_user.company_id, payload.invoice_number, exclude_id=str(purchase_id), label="Invoice Number")

    # Reverse old stock
    await adjust_parcel_stock(db, company_id=current_user.company_id, items=row.items, operation="purchase_reverse")
    await reverse_ledger_entries(db, company_id=current_user.company_id, transaction_id=str(purchase_id))

    data = payload.model_dump(exclude={"items"})
    for k, v in data.items():
        setattr(row, k, v)
    row.items.clear()
    row.items.extend([ParcelPurchaseItem(**item.model_dump()) for item in payload.items])
    _calc_totals(row)

    await _upsert_parcel_items(db, company_id=current_user.company_id, created_by_name=get_actor_name(current_user), items=payload.items, inr_rate=payload.inr_rate)
    await adjust_parcel_stock(db, company_id=current_user.company_id, items=payload.items, operation="purchase")

    # Re-post ledger
    amount = float(row.transaction_final_amount or row.total_amount or 0)
    party = row.party or "Unknown Supplier"
    if amount > 0:
        await post_ledger_entries(db, company_id=current_user.company_id, transaction_type="purchase", transaction_id=str(purchase_id), transaction_date=row.date, created_by=get_actor_name(current_user), entries=[
            {"account_name": "Purchase", "debit": amount, "credit": 0, "narration": f"Purchase {row.invoice_number} from {party}"},
            {"account_name": party, "debit": 0, "credit": amount, "narration": f"Purchase {row.invoice_number}"},
        ])

    await db.commit()
    await db.refresh(row)
    row = (await db.execute(select(ParcelPurchase).options(selectinload(ParcelPurchase.items)).where(ParcelPurchase.id == str(purchase_id)))).scalar_one()
    return ParcelPurchaseOut.model_validate(row)


@router.delete("/{purchase_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_purchase(
    purchase_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(
        select(ParcelPurchase).options(selectinload(ParcelPurchase.items))
        .where(ParcelPurchase.id == str(purchase_id), ParcelPurchase.company_id == current_user.company_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase not found")

    # Reverse stock and ledger
    await adjust_parcel_stock(db, company_id=current_user.company_id, items=row.items, operation="purchase_reverse")
    await reverse_ledger_entries(db, company_id=current_user.company_id, transaction_id=str(purchase_id))

    await db.delete(row)
    await db.commit()


@router.post("/import")
async def import_purchase_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    wb = load_workbook(filename=BytesIO(await file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty excel file")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_norm = [_norm_header(h) for h in header]
    idx = {h: i for i, h in enumerate(header_norm) if h}
    errors: list[str] = []
    added = 0
    imported_rows: list[dict] = []

    summary_required = {
        "invoice": "Invoice", "dateyyyymmdd": "Date(YYYY-MM-DD)", "type": "Type",
        "subtype": "Sub Type", "category": "Category", "party": "Party",
        "carats": "Carats", "amount": "Amount", "currency": "Currency",
        "inramt": "INR Amt", "usdamt": "USD Amt", "duedateyyyymmdd": "DueDate(YYYY-MM-DD)",
        "paymentstatus": "Payment Status",
    }
    sample_required = {
        "kapaninvno": "Kapan/Inv No", "lotno": "LotNo", "shape": "Shape",
        "size": "Size.", "color": "Color", "clarity": "Clarity", "sieve": "Sieve",
        "pcs": "Pcs", "issuecarats": "IssueCarats", "carats": "Carats",
        "purchaseamount": "PurchaseAmount",
    }

    is_summary = all(k in idx for k in summary_required)
    is_sample = all(k in idx for k in sample_required)
    if not is_summary and not is_sample:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid template schema. Found columns: {', '.join([h for h in header if h])}",
        )

    if is_summary:
        for row_no, r in enumerate(rows[1:], start=2):
            if not r or not any(v not in (None, "") for v in r):
                continue
            invoice = str(r[idx["invoice"]] or "").strip()
            if not invoice:
                errors.append(f"Row {row_no}, Column '{summary_required['invoice']}' : required")
                continue
            exists = (await db.execute(
                select(ParcelPurchase.id).where(
                    ParcelPurchase.company_id == current_user.company_id,
                    func.lower(ParcelPurchase.invoice_number) == invoice.lower(),
                )
            )).scalar_one_or_none()
            if exists:
                errors.append(f"Row {row_no}, Column '{summary_required['invoice']}' : invoice already exists")
                continue
            parsed_date = parse_date_value(r[idx["dateyyyymmdd"]])
            if not parsed_date:
                errors.append(f"Row {row_no}, Column '{summary_required['dateyyyymmdd']}' : invalid date")
                continue
            due_date_val = parse_date_value(r[idx["duedateyyyymmdd"]])
            try:
                total_carats = parse_float_value(r[idx["carats"]])
                total_amount = parse_float_value(r[idx["amount"]])
                inr_amt = parse_float_value(r[idx["inramt"]])
                usd_amt = parse_float_value(r[idx["usdamt"]])
            except Exception:
                errors.append(f"Row {row_no}: numeric parse error in Carats/Amount/INR Amt/USD Amt")
                continue
            p = ParcelPurchase(
                company_id=current_user.company_id, invoice_number=invoice, date=parsed_date,
                purchase_type=str(r[idx["type"]] or "LOCAL"), sub_type=str(r[idx["subtype"]] or ""),
                category=str(r[idx["category"]] or "Natural Diamond"), party=str(r[idx["party"]] or ""),
                total_carats=total_carats, total_amount=total_amount,
                currency=str(r[idx["currency"]] or "USD"), inr_amt=inr_amt, usd_amt=usd_amt,
                due_date=due_date_val, payment_status=str(r[idx["paymentstatus"]] or "Pending"),
                inr_final_amount=inr_amt, usd_final_amount=usd_amt, transaction_final_amount=total_amount,
                created_by_name=get_actor_name(current_user),
            )
            db.add(p)
            added += 1
    else:
        grouped: dict[str, dict] = {}
        for row_no, r in enumerate(rows[1:], start=2):
            if not r or not any(v not in (None, "") for v in r):
                continue
            invoice = str(r[idx["kapaninvno"]] or "").strip()
            if not invoice:
                errors.append(f"Row {row_no}, Column '{sample_required['kapaninvno']}' : required")
                continue
            lot_no = str(r[idx["lotno"]] or "").strip()
            if not lot_no:
                errors.append(f"Row {row_no}, Column '{sample_required['lotno']}' : required")
                continue
            shape = str(r[idx["shape"]] or "").strip()
            size = str(r[idx["size"]] or "").strip()
            if not shape:
                errors.append(f"Row {row_no}, Column '{sample_required['shape']}' : required")
                continue
            if not size:
                errors.append(f"Row {row_no}, Column '{sample_required['size']}' : required")
                continue
            try:
                pcs = int(float(r[idx["pcs"]] or 0))
                issue_carats = parse_float_value(r[idx["issuecarats"]])
                selected_carat = parse_float_value(r[idx["carats"]])
                amount = parse_float_value(r[idx["purchaseamount"]])
                sieve = str(r[idx["sieve"]] or "")
                item = ParcelPurchaseItem(
                    lot_number=lot_no, item_name=lot_no, shape=shape,
                    color=str(r[idx["color"]] or ""), clarity=str(r[idx["clarity"]] or ""),
                    size=size, sieve=sieve, issue_carats=issue_carats,
                    selected_carat=selected_carat, pcs=pcs, amount=amount,
                    rate=parse_float_value(r[idx.get("purchaserate", -1)] if "purchaserate" in idx else 0),
                )
            except Exception:
                errors.append(f"Row {row_no}: numeric parse error")
                continue
            if invoice not in grouped:
                grouped[invoice] = {"row_no": row_no, "items": []}
            grouped[invoice]["items"].append(item)

        if not errors:
            all_items: list[ParcelPurchaseItem] = []
            for invoice, data in grouped.items():
                for item in data["items"]:
                    all_items.append(item)
                    imported_rows.append({
                        "invoice_number": invoice, "lot_number": item.lot_number,
                        "item_name": item.item_name, "shape": item.shape, "color": item.color,
                        "clarity": item.clarity, "size": item.size, "sieve": item.sieve,
                        "issue_carats": item.issue_carats, "selected_carat": item.selected_carat,
                        "pcs": item.pcs, "rate": item.rate, "amount": item.amount,
                    })
            if all_items:
                await _upsert_parcel_items(db, company_id=current_user.company_id, created_by_name=get_actor_name(current_user), items=all_items, inr_rate=85)
                await adjust_parcel_stock(db, company_id=current_user.company_id, items=all_items, operation="purchase")
            added = len(imported_rows)

    if errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Import validation failed: " + "; ".join(errors[:50]),
        )
    await db.commit()
    return {"imported": added, "rows": imported_rows}
